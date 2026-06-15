# -*- coding: utf-8 -*-
"""
main.py -- Nifty 500 Swing Scanner
Run after 3:30 PM IST. Usage: python main.py
"""

import os, time, datetime
import pandas as pd
from concurrent.futures import ThreadPoolExecutor, as_completed
from symbols import get_symbols
from scanner import scan_symbol

RESULTS_DIR   = os.path.join(os.path.dirname(__file__), "results")
MAX_WORKERS   = 8
SLEEP_BETWEEN = 0.05

SETUP_COLOR = {
    "EMA Pullback":       "FFF3CD",   # yellow
    "Breakout":           "D4EDDA",   # green
    "Momentum Recovery":  "CCE5FF",   # blue
    "MACD Crossover":     "E2D9F3",   # purple
    "Supertrend Buy":     "D1ECF1",   # teal
    "52W High Breakout":  "F8D7DA",   # red/orange
    "BB Squeeze Breakout":"FCE8D5",   # orange
}
CAP_COLOR   = {"Large Cap": "D4EDDA", "Mid Cap": "FFF3CD", "Small Cap": "F8D7DA", "Unknown": "F2F2F2"}


def ensure_results_dir():
    os.makedirs(RESULTS_DIR, exist_ok=True)


def print_banner():
    print("=" * 60)
    print("   NIFTY 500 SWING SCANNER  |  Powered by Yahoo Finance")
    print("=" * 60)
    print("   Scan date : " + datetime.datetime.now().strftime("%d %b %Y  %I:%M %p"))
    print("   Setups    : EMA Pullback | Breakout | Momentum Recovery")
    print("              | MACD Crossover | Supertrend Buy | 52W High | BB Squeeze")
    print("=" * 60 + "\n")


def apply_header_style(ws, border):
    from openpyxl.styles import PatternFill, Font, Alignment
    hf = PatternFill("solid", fgColor="1F3864")
    fn = Font(bold=True, color="FFFFFF", size=10)
    al = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.fill = hf
        cell.font = fn
        cell.alignment = al
        cell.border = border
    ws.row_dimensions[1].height = 30


def save_to_excel(all_results, scan_date):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename  = "02_scan_{}.xlsx".format(timestamp)
    filepath  = os.path.join(RESULTS_DIR, filename)

    rows = []
    for res in all_results:
        rows.append({
            "Symbol":              res.get("symbol", ""),
            "Company Name":        res.get("company_name", ""),
            "Sector":              res.get("sector", ""),
            "Cap Category":        res.get("cap_category", ""),
            "Mkt Cap (INR Cr)":    res.get("market_cap_cr", ""),
            "Setup":               res.get("setup", ""),
            "Close (INR)":         res.get("close", ""),
            "Stop Loss (INR)":     res.get("stop_loss", ""),
            "Target (INR)":        res.get("target", ""),
            "RSI":                 res.get("rsi", ""),
            "Vol Ratio":           res.get("vol_ratio", ""),
            "EMA 20":              res.get("ema20", ""),
            "EMA 50":              res.get("ema50", ""),
            "EMA 200":             res.get("ema200", ""),
            "Breakout Level":      res.get("breakout_level", ""),
            "52W High":            res.get("52w_high", ""),
            "52W Low":             res.get("52w_low", ""),
            "% from 52W High":     res.get("pct_from_52w_high", ""),
            "% from 52W Low":      res.get("pct_from_52w_low", ""),
            "Beta":                res.get("beta", ""),
            "P/E (TTM)":           res.get("pe_ratio", ""),
            "Forward P/E":         res.get("forward_pe", ""),
            "P/B Ratio":           res.get("pb_ratio", ""),
            "P/S Ratio":           res.get("ps_ratio", ""),
            "PEG Ratio":           res.get("peg_ratio", ""),
            "EV/EBITDA":           res.get("ev_ebitda", ""),
            "ROE (%)":             res.get("roe", ""),
            "ROA (%)":             res.get("roa", ""),
            "Profit Margin (%)":   res.get("profit_margin", ""),
            "Oper. Margin (%)":    res.get("operating_margin", ""),
            "Revenue Growth (%)":  res.get("revenue_growth", ""),
            "Earnings Growth (%)": res.get("earnings_growth", ""),
            "Debt/Equity":         res.get("debt_to_equity", ""),
            "Current Ratio":       res.get("current_ratio", ""),
            "Quick Ratio":         res.get("quick_ratio", ""),
            "EPS TTM":             res.get("eps_ttm", ""),
            "Book Value":          res.get("book_value", ""),
            "Dividend Yield (%)":  res.get("dividend_yield", ""),
            "Data Source":         res.get("data_source", "Yahoo Finance"),
        })

    df = pd.DataFrame(rows)
    if df.empty:
        print("\n  No setups triggered today. Market may be in consolidation.\n")
        return None

    df.sort_values(["Setup", "Symbol"], inplace=True)
    df.reset_index(drop=True, inplace=True)

    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    try:
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Swing Picks")
            wb = writer.book
            ws = writer.sheets["Swing Picks"]

            apply_header_style(ws, border)

            for row_idx in range(2, len(df) + 2):
                setup = ws.cell(row=row_idx, column=6).value
                fg    = SETUP_COLOR.get(str(setup), "FFFFFF")
                fill  = PatternFill("solid", fgColor=fg)
                for ci in range(1, len(df.columns) + 1):
                    c = ws.cell(row=row_idx, column=ci)
                    c.fill = fill
                    c.border = border
                    c.alignment = Alignment(horizontal="center", vertical="center")

            cw = [10,26,16,12,14, 18,10,12,10,7,9,9,9,9,14,
                  11,11,14,13,7, 10,11,9,9,9,10, 9,9,14,14, 14,15, 11,12,11, 10,12,14]
            for i, w in enumerate(cw, 1):
                if i <= ws.max_column:
                    ws.column_dimensions[get_column_letter(i)].width = w
            ws.freeze_panes = "G2"

            # -- Fundamentals sheet --
            fcols = [
                "Symbol", "Company Name", "Sector", "Industry",
                "Cap Category", "Mkt Cap (INR Cr)",
                "P/E (TTM)", "Forward P/E", "P/B Ratio", "P/S Ratio", "PEG Ratio", "EV/EBITDA",
                "ROE (%)", "ROA (%)", "Profit Margin (%)", "Oper. Margin (%)",
                "Revenue Growth (%)", "Earnings Growth (%)",
                "Debt/Equity", "Current Ratio", "Quick Ratio",
                "EPS TTM", "Book Value", "Dividend Yield (%)",
                "52W High", "52W Low", "% from 52W High", "% from 52W Low", "Beta", "Data Source",
            ]
            fmap = {
                "Symbol": "symbol", "Company Name": "company_name",
                "Sector": "sector", "Industry": "industry",
                "Cap Category": "cap_category", "Mkt Cap (INR Cr)": "market_cap_cr",
                "P/E (TTM)": "pe_ratio", "Forward P/E": "forward_pe",
                "P/B Ratio": "pb_ratio", "P/S Ratio": "ps_ratio",
                "PEG Ratio": "peg_ratio", "EV/EBITDA": "ev_ebitda",
                "ROE (%)": "roe", "ROA (%)": "roa",
                "Profit Margin (%)": "profit_margin", "Oper. Margin (%)": "operating_margin",
                "Revenue Growth (%)": "revenue_growth", "Earnings Growth (%)": "earnings_growth",
                "Debt/Equity": "debt_to_equity", "Current Ratio": "current_ratio",
                "Quick Ratio": "quick_ratio", "EPS TTM": "eps_ttm",
                "Book Value": "book_value", "Dividend Yield (%)": "dividend_yield",
                "52W High": "52w_high", "52W Low": "52w_low",
                "% from 52W High": "pct_from_52w_high", "% from 52W Low": "pct_from_52w_low",
                "Beta": "beta",
                "Data Source": "data_source",
            }

            seen = {}
            for res in all_results:
                sym = res.get("symbol", "")
                if sym not in seen:
                    seen[sym] = res

            frows = [{col: res.get(fmap[col], "") for col in fcols} for res in seen.values()]
            df_f  = pd.DataFrame(frows, columns=fcols)
            df_f.sort_values(["Cap Category", "Symbol"], inplace=True)
            df_f.to_excel(writer, index=False, sheet_name="Fundamentals")
            ws_f = writer.sheets["Fundamentals"]
            apply_header_style(ws_f, border)

            for row_idx in range(2, len(df_f) + 2):
                cap  = ws_f.cell(row=row_idx, column=5).value
                fg   = CAP_COLOR.get(str(cap), "FFFFFF")
                fill = PatternFill("solid", fgColor=fg)
                for ci in range(1, len(fcols) + 1):
                    c = ws_f.cell(row=row_idx, column=ci)
                    c.fill = fill
                    c.border = border
                    c.alignment = Alignment(horizontal="center", vertical="center")

            fw = [10,26,18,22,12,14, 9,11,9,9,9,10, 9,9,15,14, 15,16, 11,12,11, 10,12,14, 11,11,15,13,8]
            for i, w in enumerate(fw, 1):
                ws_f.column_dimensions[get_column_letter(i)].width = w
            ws_f.freeze_panes = "C2"

            # -- Summary sheet --
            ws_s = wb.create_sheet("Summary")
            sd = df["Setup"].value_counts().reset_index()
            sd.columns = ["Setup", "Count"]
            ws_s.append(["Setup", "Count", "Scan Date"])
            for _, r in sd.iterrows():
                ws_s.append([r["Setup"], r["Count"], scan_date])
            ws_s.column_dimensions["A"].width = 22
            ws_s.column_dimensions["B"].width = 10
            ws_s.column_dimensions["C"].width = 14

            # -- Definitions sheet --
            ws_d = wb.create_sheet("Definitions")

            SEC  = PatternFill("solid", fgColor="1F3864")
            HDR  = PatternFill("solid", fgColor="2E75B6")
            SFIL = {
                "EMA Pullback":       PatternFill("solid", fgColor="FFF3CD"),
                "Breakout":           PatternFill("solid", fgColor="D4EDDA"),
                "Momentum Recovery":  PatternFill("solid", fgColor="CCE5FF"),
                "MACD Crossover":     PatternFill("solid", fgColor="E2D9F3"),
                "Supertrend Buy":     PatternFill("solid", fgColor="D1ECF1"),
                "52W High Breakout":  PatternFill("solid", fgColor="F8D7DA"),
                "BB Squeeze Breakout":PatternFill("solid", fgColor="FCE8D5"),
            }
            RULE = PatternFill("solid", fgColor="F2F2F2")
            W    = Alignment(wrap_text=True, vertical="top")
            C    = Alignment(horizontal="center", vertical="top")
            BW   = Font(bold=True, color="FFFFFF", size=11)
            BW10 = Font(bold=True, color="FFFFFF", size=10)

            defs = [
                ["SETUP DEFINITIONS", "", ""],
                ["Setup Name", "What It Means", "How to Trade It"],
                ["EMA Pullback",
                 "Stock is in uptrend (above 50 EMA & 200 EMA). Pulled back to 20 EMA, latest candle is green. RSI 35-62. Volume at or above average.",
                 "Entry: near CMP or next day open.\nStop: below 50 EMA (-2%).\nTarget: +6%.\nBest for: trending markets, mid/large caps."],
                ["Breakout",
                 "Tight consolidation (<5% range) for 5-10 days. Today closed above range top with volume 1.5x avg. RSI <72. Price above 50 EMA.",
                 "Entry: breakout close or next day open.\nStop: below consolidation low (-1%).\nTarget: +7%.\nBest for: all cap sizes."],
                ["Momentum Recovery",
                 "Long-term uptrend (above 200 EMA). RSI recently below 42, now recovering above 45. Price making higher low vs 5 days ago.",
                 "Entry: recovery candle.\nStop: below swing low (-2%).\nTarget: +5%.\nBest for: post-correction, large caps."],
                ["MACD Crossover",
                 "MACD line (12,26 EMA difference) crosses above Signal line (9-day EMA of MACD). Histogram turns positive. Price above 50 EMA. RSI 40-70.",
                 "Entry: on crossover candle or next day.\nStop: below 50 EMA (-3%).\nTarget: +6%.\nBest for: early trend confirmation, all caps."],
                ["Supertrend Buy",
                 "Supertrend indicator (ATR-based, 10,3) flips from bearish to bullish. Stock is above 200 EMA (long-term uptrend intact). Fresh signal only (crossover within 1 bar).",
                 "Entry: on Supertrend flip candle.\nStop: below Supertrend line (-1%).\nTarget: +7%.\nBest for: trending stocks, ride the full move."],
                ["52W High Breakout",
                 "Stock is at or above its 52-week high with volume 1.3x average. Price above 50 EMA. RSI <80. Breakout to new highs often leads to strong continuation.",
                 "Entry: on breakout candle close.\nStop: -5% from entry (wide, to avoid shakeout).\nTarget: +10%.\nBest for: strong trending stocks, momentum plays."],
                ["BB Squeeze Breakout",
                 "Bollinger Bands narrowed (squeeze = low volatility) over recent days, then price breaks above the upper band with expanding bandwidth and volume. Squeeze often precedes explosive moves.",
                 "Entry: on upper band breakout candle.\nStop: below Bollinger midline (-1%).\nTarget: +8%.\nBest for: coiled stocks about to move, all caps."],
                ["", "", ""],
                ["COLUMN GLOSSARY", "", ""],
                ["Column", "Definition", "Benchmark"],
                ["Cap Category",        "SEBI: Large Cap >20,000 Cr | Mid 5,000-20,000 Cr | Small <5,000 Cr", ""],
                ["P/E (TTM)",           "Price / Trailing 12m EPS. What you pay per rupee of profit.", "Fair: 15-25 | Expensive: >40"],
                ["Forward P/E",         "P/E on next-year estimated earnings. Lower than TTM = growth expected.", "Lower is better"],
                ["P/B Ratio",           "Price / Book Value per share.", "Good: <3 | Value: <1"],
                ["P/S Ratio",           "Price / Revenue per share.", "Good: <3"],
                ["PEG Ratio",           "P/E / Earnings growth. Adjusts valuation for growth.", "Buy: <1 | Fair: 1-2"],
                ["EV/EBITDA",           "Enterprise value / Operating profit.", "Good: <10 | Fair: 10-15"],
                ["ROE (%)",             "Profit per rupee of shareholder equity.", "Strong: >15% | Excellent: >25%"],
                ["ROA (%)",             "Profit per rupee of total assets.", "Strong: >10%"],
                ["Profit Margin (%)",   "Net profit as % of revenue.", "Good: >10% | Excellent: >20%"],
                ["Oper. Margin (%)",    "Operating profit as % of revenue.", "Good: >15%"],
                ["Revenue Growth (%)",  "YoY revenue growth.", "Good: >10% | Strong: >20%"],
                ["Earnings Growth (%)", "YoY net profit growth.", "Good: >15% | Strong: >25%"],
                ["Debt/Equity",         "Total debt / Equity.", "Safe: <0.5 | Risky: >2"],
                ["Current Ratio",       "Current assets / Current liabilities.", "Healthy: >1.5"],
                ["Quick Ratio",         "Like current ratio but excludes inventory.", "Healthy: >1"],
                ["EPS TTM",             "Earnings per share (trailing 12 months).", "Rising trend is key"],
                ["Book Value",          "Net asset value per share.", "Near book = potentially undervalued"],
                ["Dividend Yield (%)",  "Annual dividend / Price x 100.", "Income: >2%"],
                ["% from 52W High",     "How far below 52-week peak (negative = below).", "Sweet spot: -5% to -20%"],
                ["Beta",                "Volatility vs Nifty. >1 = more volatile, <1 = calmer.", "Swing: 0.8-1.4"],
                ["Vol Ratio",           "Today volume / 20-day avg volume.", "Signal: >1.5"],
                ["MACD Val / Signal",   "MACD line and Signal line values. Positive MACD = bullish momentum.", "MACD > 0 = uptrend"],
                ["Supertrend Support",  "Supertrend lower band level = dynamic stop loss for Supertrend Buy setup.", "Stay above this"],
                ["BB Upper",            "Bollinger Band upper boundary. Price above this = breakout.", ""],
                ["", "", ""],
                ["RISK MANAGEMENT RULES", "", ""],
                ["Rule", "Details", ""],
                ["Position sizing", "Qty = (Capital x 1%) / (Entry - Stop). Max 2% risk per trade.", ""],
                ["Max open trades", "No more than 5-6 positions at a time.", ""],
                ["Stop loss",       "Place stop immediately after entry. Never average down on losers.", ""],
                ["Market filter",   "Only go long when Nifty 50 is above its 20 EMA.", ""],
                ["VIX filter",      "Skip new entries when India VIX > 20.", ""],
                ["Stale exit",      "Exit if stock does not move within 3 trading days.", ""],
                ["Setup priority",  "52W High Breakout + BB Squeeze = highest conviction. MACD + Supertrend = trend following. EMA Pullback = lowest risk entry.", ""],
            ]

            SECTIONS = {"SETUP DEFINITIONS", "COLUMN GLOSSARY", "RISK MANAGEMENT RULES"}
            COL_HDRS = {"Setup Name", "Column", "Rule"}
            SETUPS   = {"EMA Pullback", "Breakout", "Momentum Recovery", "MACD Crossover", "Supertrend Buy", "52W High Breakout", "BB Squeeze Breakout"}

            for row_data in defs:
                ws_d.append(row_data)
                r = ws_d.max_row
                v = str(row_data[0]) if row_data[0] else ""
                if v in SECTIONS:
                    ws_d.merge_cells("A{}:C{}".format(r, r))
                    ws_d["A{}".format(r)].fill = SEC
                    ws_d["A{}".format(r)].font = BW
                    ws_d["A{}".format(r)].alignment = Alignment(horizontal="center", vertical="center")
                    ws_d.row_dimensions[r].height = 22
                elif v in COL_HDRS:
                    for col in ["A", "B", "C"]:
                        ws_d["{}{}".format(col, r)].fill = HDR
                        ws_d["{}{}".format(col, r)].font = BW10
                        ws_d["{}{}".format(col, r)].alignment = C
                    ws_d.row_dimensions[r].height = 18
                elif v in SETUPS:
                    for col in ["A", "B", "C"]:
                        ws_d["{}{}".format(col, r)].fill = SFIL.get(v)
                        ws_d["{}{}".format(col, r)].font = Font(bold=(col == "A"), size=10)
                        ws_d["{}{}".format(col, r)].alignment = W
                    ws_d.row_dimensions[r].height = 75
                elif v:
                    for col in ["A", "B", "C"]:
                        ws_d["{}{}".format(col, r)].fill = RULE
                        ws_d["{}{}".format(col, r)].alignment = W
                        ws_d["{}{}".format(col, r)].font = Font(size=10)
                    ws_d.row_dimensions[r].height = 38

            ws_d.column_dimensions["A"].width = 22
            ws_d.column_dimensions["B"].width = 65
            ws_d.column_dimensions["C"].width = 26

            wb.move_sheet("Definitions", offset=-wb.index(wb["Definitions"]))

    except PermissionError:
        print("\n  ERROR: Cannot write Excel file.")
        print("  >> Close the previous scan result in Excel, then re-run.\n")
        return None
    except Exception as e:
        print("\n  ERROR saving Excel: {}\n".format(e))
        return None

    print("\n  Results saved --> " + filepath)
    return filepath


def main():
    print_banner()
    ensure_results_dir()

    symbols = get_symbols()
    total   = len(symbols)
    print("  Scanning {} symbols...\n".format(total))

    all_results, scanned, triggered, failed = [], 0, 0, 0

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_to_sym = {executor.submit(scan_symbol, sym): sym for sym in symbols}
        for future in as_completed(future_to_sym):
            scanned += 1
            try:
                results = future.result()
                if results:
                    all_results.extend(results)
                    triggered += len(results)
                    for r in results:
                        print("  + {:<18} {:<22} CMP: {:<10}  RSI: {}".format(
                            r["symbol"], r["setup"], r["close"], r["rsi"]))
            except Exception:
                failed += 1

            if scanned % 50 == 0:
                pct = scanned / total * 100
                print("\n  [{}/{}  {:.0f}%]  Setups so far: {}\n".format(
                    scanned, total, pct, triggered))
            time.sleep(SLEEP_BETWEEN)

    scan_date = datetime.datetime.now().strftime("%Y-%m-%d")
    print("\n" + "=" * 60)
    print("  SCAN COMPLETE | Scanned: {} | Setups: {} | Errors: {}".format(
        scanned, triggered, failed))
    print("=" * 60)
    save_to_excel(all_results, scan_date)
    print("\n  Tip: Review each stock on TradingView before placing orders.")
    print("  Max 1-2% capital risk per trade.\n")


if __name__ == "__main__":
    main()
