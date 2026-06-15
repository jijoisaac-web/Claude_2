# -*- coding: utf-8 -*-
"""
trade_planner.py -- Swing Trade Position Planner
-------------------------------------------------
Capital  : INR 20,00,000
Broker   : Zerodha NRE account (delivery-based swing trades)
Usage    : python trade_planner.py

Reads the latest 02_scan_*.xlsx from /results and generates a
complete trade plan with:
  - Position sizing (1% risk per trade)
  - Zerodha NRE brokerage + all statutory charges
  - Net profit/loss after charges
  - Actual Risk:Reward after charges
  - Capital allocation summary
"""

import os
import glob
import datetime
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------
CAPITAL          = 2_000_000   # INR 20 lakhs
RISK_PCT         = 1.0         # % of capital to risk per trade
MAX_POSITIONS    = 6           # max simultaneous open trades
RESULTS_DIR      = os.path.join(os.path.dirname(__file__), "results")

# Priority order — stocks with multiple setups ranked higher
PRIORITY_STOCKS  = [
    "TITAN", "EQUITASBNK", "GREENPLY", "VARROC", "RBLBANK",
    "GICRE", "VIJAYA", "IPCALAB", "POWERMECH", "BHARATFORG",
    "SUPREMEIND", "BAJAJ-AUTO", "DIVISLAB", "MARICO", "FORTIS",
    "SONACOMS", "JBCHEPHARM", "ABCAPITAL", "BSE", "GRSE",
    "APTUS", "GMRAIRPORT", "IIFL", "SBIN", "GAIL",
]

# ---------------------------------------------------------------
# Zerodha NRE Account — Charge Structure (Delivery / CNC)
# NRI delivery trades are NOT free — ₹200 per executed order
# ---------------------------------------------------------------

def zerodha_nre_charges(buy_value, sell_value):
    """
    Calculate all-in charges for a round-trip delivery trade.

    Zerodha NRE (NRI) — Equity Delivery:
      Brokerage     : INR 200 per executed order (buy + sell = INR 400)
      STT           : 0.1% on buy value + 0.1% on sell value
      NSE Txn charge: 0.00322% on total turnover (buy + sell)
      GST           : 18% on (brokerage + exchange charges)
      SEBI charge   : 0.0001% on total turnover
      Stamp duty    : 0.015% on buy value only
    """
    turnover   = buy_value + sell_value

    brokerage  = 400.0                                # ₹200 buy + ₹200 sell
    stt        = (buy_value * 0.001) + (sell_value * 0.001)
    exch_chrg  = turnover * 0.0000322
    sebi       = turnover * 0.000001
    stamp      = buy_value * 0.00015
    gst        = (brokerage + exch_chrg) * 0.18

    total      = brokerage + stt + exch_chrg + sebi + stamp + gst

    return {
        "brokerage":   round(brokerage, 2),
        "stt":         round(stt, 2),
        "exch_chrg":   round(exch_chrg, 2),
        "sebi":        round(sebi, 2),
        "stamp":       round(stamp, 2),
        "gst":         round(gst, 2),
        "total":       round(total, 2),
    }


# ---------------------------------------------------------------
# Position sizing
# ---------------------------------------------------------------

# Max single position = 25% of capital (concentration limit)
MAX_POSITION_PCT = 15.0   # 15% cap → 6 trades × 15% = 90% deployed, 10% reserve

def position_size(capital, risk_pct, entry, stop):
    """
    Qty = (Capital x Risk%) / (Entry - Stop)
    Capped at MAX_POSITION_PCT% of capital per trade.
    """
    risk_amount    = capital * (risk_pct / 100)
    risk_per_share = entry - stop
    if risk_per_share <= 0:
        return 0, 0
    qty       = int(risk_amount / risk_per_share)
    pos_value = qty * entry
    # Apply concentration cap
    max_value = capital * (MAX_POSITION_PCT / 100)
    if pos_value > max_value:
        qty       = int(max_value / entry)
        pos_value = qty * entry
    return qty, round(pos_value, 2)


# ---------------------------------------------------------------
# Load latest scan results
# ---------------------------------------------------------------

def load_latest_scan():
    files = sorted(glob.glob(os.path.join(RESULTS_DIR, "02_scan_*.xlsx")))
    if not files:
        print("  No scan files found. Run main.py first.")
        return None
    latest = files[-1]
    print("  Loading: " + os.path.basename(latest))
    df = pd.read_excel(latest, sheet_name="Swing Picks")
    return df


# ---------------------------------------------------------------
# Build trade plan
# ---------------------------------------------------------------

def build_trade_plan(df):
    # Deduplicate: if same stock has multiple setups, pick highest conviction
    # (most setups = first in priority list)
    seen = {}
    for _, row in df.iterrows():
        sym   = str(row.get("Symbol", "")).strip()
        setup = str(row.get("Setup", "")).strip()
        if sym not in seen:
            seen[sym] = row
        else:
            # If same stock appears again with a different setup, append setup name
            existing = seen[sym]
            existing_setup = str(existing.get("Setup", ""))
            if setup not in existing_setup:
                seen[sym] = existing.copy()
                seen[sym]["Setup"] = existing_setup + " + " + setup

    # Sort by priority list
    def priority_rank(sym):
        try:
            return PRIORITY_STOCKS.index(sym)
        except ValueError:
            return 999

    sorted_stocks = sorted(seen.values(), key=lambda r: priority_rank(str(r.get("Symbol", ""))))

    # Build plan rows — take top MAX_POSITIONS
    plan_rows = []
    capital_deployed = 0
    capital_remaining = CAPITAL

    for row in sorted_stocks[:MAX_POSITIONS]:
        sym    = str(row.get("Symbol", "")).strip()
        setup  = str(row.get("Setup", "")).strip()
        entry  = float(row.get("Close (INR)", 0) or 0)
        stop   = float(row.get("Stop Loss (INR)", 0) or 0)
        target = float(row.get("Target (INR)", 0) or 0)
        rsi    = row.get("RSI", "")
        vol    = row.get("Vol Ratio", "")
        cap_cat = str(row.get("Cap Category", "") or "")
        sector  = str(row.get("Sector", "") or "")
        pe      = row.get("P/E (TTM)", "")
        roe     = row.get("ROE (%)", "")

        if entry <= 0 or stop <= 0 or target <= 0:
            continue

        qty, pos_value = position_size(CAPITAL, RISK_PCT, entry, stop)
        if qty == 0:
            continue

        buy_value  = qty * entry
        sell_value = qty * target   # charges at target (best case)
        stop_sell  = qty * stop     # charges at stop (worst case)

        charges_win  = zerodha_nre_charges(buy_value, sell_value)
        charges_loss = zerodha_nre_charges(buy_value, stop_sell)

        gross_profit = (target - entry) * qty
        gross_loss   = (entry - stop)   * qty

        net_profit   = gross_profit - charges_win["total"]
        net_loss     = gross_loss   + charges_loss["total"]   # loss + charges

        rr_gross     = gross_profit / gross_loss if gross_loss > 0 else 0
        rr_net       = net_profit / net_loss     if net_loss   > 0 else 0

        charges_pct  = charges_win["total"] / buy_value * 100

        capital_deployed += buy_value

        plan_rows.append({
            "Symbol":             sym,
            "Sector":             sector,
            "Cap":                cap_cat,
            "Setup(s)":           setup,
            "Entry (INR)":        entry,
            "Stop (INR)":         stop,
            "Target (INR)":       target,
            "Qty":                qty,
            "Investment (INR)":   round(buy_value, 0),
            "% of Capital":       round(buy_value / CAPITAL * 100, 1),
            "Risk (INR)":         round(gross_loss, 0),
            "Reward (INR)":       round(gross_profit, 0),
            "Gross R:R":          round(rr_gross, 2),
            "Total Charges (INR)":round(charges_win["total"], 0),
            "Charges %":          round(charges_pct, 3),
            "Net Profit (INR)":   round(net_profit, 0),
            "Net Loss (INR)":     round(net_loss, 0),
            "Net R:R":            round(rr_net, 2),
            "RSI":                rsi,
            "Vol Ratio":          vol,
            "P/E":                pe,
            "ROE (%)":            roe,
            # Charge breakdown
            "Brokerage (INR)":    charges_win["brokerage"],
            "STT (INR)":          charges_win["stt"],
            "Exch Chrg (INR)":    charges_win["exch_chrg"],
            "GST (INR)":          charges_win["gst"],
            "Stamp Duty (INR)":   charges_win["stamp"],
            "SEBI (INR)":         charges_win["sebi"],
        })

    return plan_rows, capital_deployed


# ---------------------------------------------------------------
# Save to Excel
# ---------------------------------------------------------------

def save_plan(plan_rows, capital_deployed):
    if not plan_rows:
        print("  No valid trades to plan.")
        return

    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename  = "02_trade_plan_{}.xlsx".format(timestamp)
    filepath  = os.path.join(RESULTS_DIR, filename)

    df = pd.DataFrame(plan_rows)

    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    COLORS = {
        "header":     "1F3864",
        "pos_main":   "E8F4FD",
        "charges":    "FFF3CD",
        "summary":    "D4EDDA",
        "alert":      "F8D7DA",
    }

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:

        # ── Sheet 1: Trade Plan ──────────────────────────────────
        main_cols = [
            "Symbol", "Sector", "Cap", "Setup(s)",
            "Entry (INR)", "Stop (INR)", "Target (INR)", "Qty",
            "Investment (INR)", "% of Capital",
            "Risk (INR)", "Reward (INR)", "Gross R:R",
            "Total Charges (INR)", "Charges %",
            "Net Profit (INR)", "Net Loss (INR)", "Net R:R",
            "RSI", "Vol Ratio", "P/E", "ROE (%)",
        ]
        df[main_cols].to_excel(writer, index=False, sheet_name="Trade Plan")
        wb = writer.book
        ws = writer.sheets["Trade Plan"]

        hf  = PatternFill("solid", fgColor=COLORS["header"])
        hfn = Font(bold=True, color="FFFFFF", size=10)
        for cell in ws[1]:
            cell.fill = hf; cell.font = hfn
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[1].height = 32

        for row_idx in range(2, len(df) + 2):
            rr = ws.cell(row=row_idx, column=18).value   # Net R:R col
            try:
                fg = "D4EDDA" if float(rr) >= 1.5 else ("FFF3CD" if float(rr) >= 1.0 else "F8D7DA")
            except Exception:
                fg = "F2F2F2"
            fill = PatternFill("solid", fgColor=fg)
            for ci in range(1, len(main_cols) + 1):
                c = ws.cell(row=row_idx, column=ci)
                c.fill = fill; c.border = border
                c.alignment = Alignment(horizontal="center", vertical="center")

        cw = [12,18,10,26, 11,11,11,7, 16,12, 12,13,10, 18,10, 16,14,9, 6,9,7,8]
        for i, w in enumerate(cw, 1):
            if i <= ws.max_column:
                ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "E2"

        # ── Sheet 2: Charge Breakdown ────────────────────────────
        chg_cols = [
            "Symbol", "Qty", "Investment (INR)",
            "Brokerage (INR)", "STT (INR)", "Exch Chrg (INR)",
            "GST (INR)", "Stamp Duty (INR)", "SEBI (INR)",
            "Total Charges (INR)", "Charges %",
            "Gross R:R", "Net R:R",
        ]
        df[chg_cols].to_excel(writer, index=False, sheet_name="Charge Breakdown")
        ws_c = writer.sheets["Charge Breakdown"]
        for cell in ws_c[1]:
            cell.fill = PatternFill("solid", fgColor="2E75B6")
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        ws_c.row_dimensions[1].height = 28
        for row_idx in range(2, len(df) + 2):
            for ci in range(1, len(chg_cols) + 1):
                c = ws_c.cell(row=row_idx, column=ci)
                c.fill = PatternFill("solid", fgColor="FFF3CD")
                c.border = border
                c.alignment = Alignment(horizontal="center", vertical="center")
        for i, w in enumerate([12,7,16, 14,12,14, 12,14,12, 18,10, 10,9], 1):
            ws_c.column_dimensions[get_column_letter(i)].width = w

        # ── Sheet 3: Capital Summary ─────────────────────────────
        ws_sum = wb.create_sheet("Capital Summary")

        total_risk      = sum(r["Risk (INR)"]    for r in plan_rows)
        total_reward    = sum(r["Reward (INR)"]  for r in plan_rows)
        total_charges   = sum(r["Total Charges (INR)"] for r in plan_rows)
        total_net_profit = sum(r["Net Profit (INR)"] for r in plan_rows)
        total_net_loss   = sum(r["Net Loss (INR)"]   for r in plan_rows)
        avg_rr          = round(sum(r["Net R:R"] for r in plan_rows) / len(plan_rows), 2)
        capital_remaining = CAPITAL - capital_deployed
        deployed_pct    = round(capital_deployed / CAPITAL * 100, 1)

        summary = [
            ["CAPITAL SUMMARY", ""],
            ["Total Capital",              "INR {:,.0f}".format(CAPITAL)],
            ["Capital Deployed",           "INR {:,.0f}  ({:.1f}%)".format(capital_deployed, deployed_pct)],
            ["Capital in Reserve",         "INR {:,.0f}  ({:.1f}%)".format(capital_remaining, 100 - deployed_pct)],
            ["Number of Trades",           len(plan_rows)],
            ["", ""],
            ["RISK & REWARD (ALL TRADES)", ""],
            ["Total Risk (if all stop out)","INR {:,.0f}".format(total_risk)],
            ["Risk as % of Capital",        "{:.2f}%".format(total_risk / CAPITAL * 100)],
            ["Total Reward (if all target)","INR {:,.0f}".format(total_reward)],
            ["Reward as % of Capital",      "{:.2f}%".format(total_reward / CAPITAL * 100)],
            ["Portfolio Gross R:R",         round(total_reward / total_risk, 2) if total_risk else 0],
            ["", ""],
            ["AFTER CHARGES", ""],
            ["Total Charges (win scenario)", "INR {:,.0f}".format(total_charges)],
            ["Net Profit (all targets hit)", "INR {:,.0f}".format(total_net_profit)],
            ["Net Loss (all stops hit)",     "INR {:,.0f}".format(total_net_loss)],
            ["Avg Net R:R per trade",        avg_rr],
            ["Portfolio Net R:R",            round(total_net_profit / total_net_loss, 2) if total_net_loss else 0],
            ["", ""],
            ["ZERODHA NRE CHARGES STRUCTURE", ""],
            ["Brokerage",    "INR 200 per order (buy) + INR 200 (sell) = INR 400 per trade"],
            ["STT",          "0.1% on buy value + 0.1% on sell value"],
            ["NSE Exch Chrg","0.00322% on total turnover (buy + sell)"],
            ["GST",          "18% on brokerage + exchange charges"],
            ["SEBI",         "0.0001% on total turnover"],
            ["Stamp Duty",   "0.015% on buy value only"],
            ["Intraday",     "NOT permitted on NRE account"],
            ["F&O",          "NOT permitted on NRE account"],
            ["", ""],
            ["POSITION SIZING RULES", ""],
            ["Risk per trade",    "1% of total capital = INR {:,.0f}".format(CAPITAL * 0.01)],
            ["Max positions",     "{} simultaneous trades".format(MAX_POSITIONS)],
            ["Stop placement",    "Technical stop (EMA / consolidation low) — not arbitrary %"],
            ["Exit discipline",   "Exit at target OR stop — no holding beyond plan"],
            ["", ""],
            ["BREAK-EVEN ANALYSIS", ""],
            ["Win rate needed to break even", "{:.0f}%".format(
                100 / (1 + avg_rr) if avg_rr > 0 else 50)],
            ["If win rate = 40%",  "Expected P&L = INR {:,.0f}".format(
                0.4 * total_net_profit - 0.6 * total_net_loss)],
            ["If win rate = 50%",  "Expected P&L = INR {:,.0f}".format(
                0.5 * total_net_profit - 0.5 * total_net_loss)],
            ["If win rate = 60%",  "Expected P&L = INR {:,.0f}".format(
                0.6 * total_net_profit - 0.4 * total_net_loss)],
        ]

        sec_fill  = PatternFill("solid", fgColor="1F3864")
        sec_font  = Font(bold=True, color="FFFFFF", size=11)
        val_fill  = PatternFill("solid", fgColor="F2F2F2")
        hi_fill   = PatternFill("solid", fgColor="D4EDDA")
        warn_fill = PatternFill("solid", fgColor="F8D7DA")

        SECTIONS = {
            "CAPITAL SUMMARY", "RISK & REWARD (ALL TRADES)", "AFTER CHARGES",
            "ZERODHA NRE CHARGES STRUCTURE", "POSITION SIZING RULES", "BREAK-EVEN ANALYSIS"
        }

        for row_data in summary:
            ws_sum.append(row_data)
            r = ws_sum.max_row
            v = str(row_data[0])

            if v in SECTIONS:
                ws_sum.merge_cells("A{}:B{}".format(r, r))
                ws_sum["A{}".format(r)].fill = sec_fill
                ws_sum["A{}".format(r)].font = sec_font
                ws_sum["A{}".format(r)].alignment = Alignment(horizontal="left", vertical="center")
                ws_sum.row_dimensions[r].height = 22
            elif v:
                is_good = any(x in v for x in ["Net Profit", "Reward", "60%", "50%"])
                is_warn = any(x in v for x in ["Net Loss", "stop out", "NOT"])
                fill = hi_fill if is_good else (warn_fill if is_warn else val_fill)
                for col in ["A", "B"]:
                    ws_sum["{}{}".format(col, r)].fill = fill
                    ws_sum["{}{}".format(col, r)].font = Font(size=10, bold=(col == "A"))
                    ws_sum["{}{}".format(col, r)].alignment = Alignment(
                        horizontal="left" if col == "A" else "right",
                        vertical="center", wrap_text=True)
                    ws_sum["{}{}".format(col, r)].border = border
                ws_sum.row_dimensions[r].height = 20

        ws_sum.column_dimensions["A"].width = 35
        ws_sum.column_dimensions["B"].width = 55

        # Sheet order
        wb.move_sheet("Capital Summary", offset=-wb.index(wb["Capital Summary"]))

    print("\n  Trade plan saved --> " + filepath)
    return filepath


# ---------------------------------------------------------------
# Main
# ---------------------------------------------------------------

def main():
    print("=" * 60)
    print("   SWING TRADE PLANNER  |  Zerodha NRE  |  INR 20 Lakhs")
    print("=" * 60)
    print("   Capital    : INR {:,.0f}".format(CAPITAL))
    print("   Risk/trade : {:.0f}% = INR {:,.0f}".format(RISK_PCT, CAPITAL * RISK_PCT / 100))
    print("   Max trades : {}".format(MAX_POSITIONS))
    print("   Broker     : Zerodha NRE (INR 400 brokerage/trade)")
    print("=" * 60 + "\n")

    df = load_latest_scan()
    if df is None:
        return

    plan_rows, capital_deployed = build_trade_plan(df)
    if not plan_rows:
        print("  No trades to plan.")
        return

    print("\n  {:<16} {:<26} {:>8} {:>12} {:>12} {:>12} {:>10} {:>8}".format(
        "Symbol", "Setup", "Qty", "Invest(INR)", "Charges", "Net Profit", "Net Loss", "Net R:R"))
    print("  " + "-" * 108)
    for r in plan_rows:
        print("  {:<16} {:<26} {:>8,} {:>12,.0f} {:>12,.0f} {:>12,.0f} {:>10,.0f} {:>8.2f}".format(
            r["Symbol"], r["Setup(s)"][:25], r["Qty"],
            r["Investment (INR)"], r["Total Charges (INR)"],
            r["Net Profit (INR)"], r["Net Loss (INR)"], r["Net R:R"]))

    total_invest  = sum(r["Investment (INR)"]   for r in plan_rows)
    total_charges = sum(r["Total Charges (INR)"] for r in plan_rows)
    total_profit  = sum(r["Net Profit (INR)"]    for r in plan_rows)
    total_loss    = sum(r["Net Loss (INR)"]      for r in plan_rows)

    print("  " + "-" * 108)
    print("  {:<42} {:>12,.0f} {:>12,.0f} {:>12,.0f} {:>10,.0f}".format(
        "TOTAL ({} trades)".format(len(plan_rows)),
        total_invest, total_charges, total_profit, total_loss))
    print("\n  Capital deployed : INR {:,.0f}  ({:.1f}% of total)".format(
        total_invest, total_invest / CAPITAL * 100))
    print("  Capital reserve  : INR {:,.0f}  ({:.1f}% of total)".format(
        CAPITAL - total_invest, (CAPITAL - total_invest) / CAPITAL * 100))
    print("  Total charges    : INR {:,.0f}  (round trip, win scenario)".format(total_charges))

    save_plan(plan_rows, capital_deployed)
    print("\n  Open trade_plan_*.xlsx in the results folder.\n")


if __name__ == "__main__":
    main()
