# -*- coding: utf-8 -*-
"""
report_generator.py — Excel Report Generator
─────────────────────────────────────────────
Sheets:
  1. Portfolio Plan     — Final selected trades with full risk metrics
  2. All Signals        — Every signal from every strategy (unfiltered)
  3. Mean Reversion     — Strategy-specific detail
  4. Momentum           — Strategy-specific detail
  5. Factor Model       — Top 20 ranked stocks
  6. Pairs Trading      — Stat arb opportunities
  7. Sector Rotation    — Sector rankings + leading stocks
  8. PEAD               — Earnings gap plays
  9. Charge Breakdown   — Zerodha NRE charges per trade
 10. Risk Summary       — Portfolio-level metrics & capital allocation
"""

import os
import datetime
import pandas as pd
import numpy as np
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils  import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from config import CAPITAL, RESULTS_DIR

os.makedirs(RESULTS_DIR, exist_ok=True)

# ── Style constants ───────────────────────────────────────────────────────────
C = {
    "hdr_dark":    "1F3864",
    "hdr_blue":    "2E75B6",
    "hdr_green":   "375623",
    "hdr_purple":  "4B0082",
    "hdr_teal":    "1F5C6B",
    "hdr_orange":  "8B4513",
    "green_light": "D4EDDA",
    "red_light":   "F8D7DA",
    "yellow_light":"FFF3CD",
    "blue_light":  "D6E4F0",
    "purple_light":"E8D5F5",
    "teal_light":  "D0F0F5",
    "white":       "FFFFFF",
    "gray_light":  "F2F2F2",
    "gray_mid":    "D9D9D9",
}

thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def _hdr_fill(color_key): return PatternFill("solid", fgColor=C[color_key])
def _cell_fill(color_key): return PatternFill("solid", fgColor=C[color_key])
def _hdr_font(size=10): return Font(bold=True, color="FFFFFF", size=size)
def _body_font(bold=False, size=10): return Font(bold=bold, size=size)
def _center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def _left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)
def _right():  return Alignment(horizontal="right",  vertical="center")


def _style_header_row(ws, row: int, color_key: str, height: int = 30):
    for cell in ws[row]:
        cell.fill      = _hdr_fill(color_key)
        cell.font      = _hdr_font()
        cell.alignment = _center()
        cell.border    = BORDER
    ws.row_dimensions[row].height = height


def _style_data_rows(ws, start_row: int, end_row: int, fill_color: str,
                     alt_color: str = None):
    for r in range(start_row, end_row + 1):
        bg = fill_color if (alt_color is None or r % 2 == 0) else alt_color
        fill = PatternFill("solid", fgColor=bg)
        for cell in ws[r]:
            cell.fill      = fill
            cell.border    = BORDER
            cell.alignment = _center()
        ws.row_dimensions[r].height = 18


def _set_col_widths(ws, widths: list):
    for i, w in enumerate(widths, 1):
        if i <= ws.max_column:
            ws.column_dimensions[get_column_letter(i)].width = w


# ── Sheet builders ────────────────────────────────────────────────────────────

def _sheet_portfolio(wb, portfolio: list):
    ws = wb.create_sheet("📊 Portfolio Plan")
    if not portfolio:
        ws.append(["No trades selected"])
        return

    cols = [
        "Symbol", "Strategies", "Setup", "Direction",
        "Entry", "Stop", "Target", "Gross R:R", "Net R:R",
        "Qty", "Investment (₹)", "% Capital",
        "Risk (₹)", "Reward (₹)", "Risk % Capital",
        "Charges (₹)", "Charges %",
        "Net Profit (₹)", "Net Loss (₹)", "BE Win Rate %",
        "RSI", "ADV (Cr)", "Sector", "Score", "Conviction"
    ]
    ws.append(cols)
    _style_header_row(ws, 1, "hdr_dark", 32)

    for t in portfolio:
        net_rr = t.get("net_rr", 0) or 0
        ws.append([
            t.get("symbol"),
            t.get("strategies", ""),
            t.get("all_setups", t.get("setup", ""))[:60],
            t.get("direction"),
            t.get("entry"),    t.get("stop"),    t.get("target"),
            t.get("gross_rr"), net_rr,
            t.get("qty"),
            t.get("investment"), t.get("pct_capital"),
            t.get("risk_amount"), t.get("reward_amount"), t.get("risk_pct_cap"),
            t.get("charges_win"), t.get("charges_pct"),
            t.get("net_profit"), t.get("net_loss"), t.get("be_win_rate"),
            t.get("rsi"), t.get("adv_cr"),
            t.get("sector"), t.get("final_score"), t.get("conviction"),
        ])

    for r in range(2, len(portfolio) + 2):
        rr = ws.cell(row=r, column=9).value or 0
        bg = C["green_light"] if rr >= 1.5 else (C["yellow_light"] if rr >= 1.0 else C["red_light"])
        fill = PatternFill("solid", fgColor=bg)
        for ci in range(1, len(cols) + 1):
            c = ws.cell(row=r, column=ci)
            c.fill = fill; c.border = BORDER; c.alignment = _center()
        ws.row_dimensions[r].height = 18

    _set_col_widths(ws, [12,28,40,10, 9,9,9,9,9, 7,14,9, 11,11,12, 12,9, 14,12,12, 6,8,14,8,8])
    ws.freeze_panes = "E2"


def _sheet_all_signals(wb, scored: list):
    ws = wb.create_sheet("🔎 All Signals")
    if not scored:
        ws.append(["No signals"]); return

    cols = ["Symbol","Strategy","Strategies Hit","Setup","Direction",
            "Entry","Stop","Target","R:R","Conviction","Score",
            "RSI","ADV(Cr)","Sector"]
    ws.append(cols)
    _style_header_row(ws, 1, "hdr_blue")

    for s in scored[:100]:   # top 100
        ws.append([
            s.get("symbol"), s.get("strategy",""), s.get("strategies",""),
            s.get("all_setups", s.get("setup",""))[:50],
            s.get("direction"),
            s.get("entry"), s.get("stop"), s.get("target"), s.get("rr"),
            s.get("conviction"), s.get("final_score"),
            s.get("rsi"), s.get("adv_cr"), s.get("sector",""),
        ])
    _style_data_rows(ws, 2, len(scored[:100]) + 1, C["blue_light"], C["white"])
    _set_col_widths(ws, [12,18,28,50,10, 9,9,9,8,10,8, 6,8,14])
    ws.freeze_panes = "E2"


def _sheet_strategy(wb, name: str, signals: list, cols_def: list, header_color: str,
                    fill_color: str):
    ws = wb.create_sheet(name)
    if not signals:
        ws.append([f"No signals for {name}"]); return

    col_names = [c[0] for c in cols_def]
    col_keys  = [c[1] for c in cols_def]
    ws.append(col_names)
    _style_header_row(ws, 1, header_color)

    for s in signals:
        ws.append([s.get(k) for k in col_keys])

    _style_data_rows(ws, 2, len(signals) + 1, fill_color, C["white"])
    widths = [c[2] if len(c) > 2 else 12 for c in cols_def]
    _set_col_widths(ws, widths)
    ws.freeze_panes = "D2"


def _sheet_sector_rotation(wb, sector_info: list, sector_picks: list):
    ws = wb.create_sheet("🔄 Sector Rotation")

    # Section 1: Sector Rankings
    ws.append(["SECTOR RANKINGS"])
    ws.merge_cells(f"A1:H1")
    ws["A1"].fill = _hdr_fill("hdr_teal"); ws["A1"].font = _hdr_font(12)
    ws["A1"].alignment = _left()
    ws.row_dimensions[1].height = 24

    ws.append(["Sector", "Rank", "Signal", "30D Return %", "90D Return %",
               "Momentum Score", "# Stocks", ""])
    _style_header_row(ws, 2, "hdr_teal")

    for s in sector_info:
        sig  = s.get("signal", "")
        fill_c = C["green_light"] if sig == "OVERWEIGHT" else (
                 C["red_light"]   if sig == "UNDERWEIGHT" else C["gray_light"])
        row = ws.max_row + 1
        ws.append([s.get("sector"), s.get("rank"), sig,
                   s.get("ret_30d"), s.get("ret_90d"), s.get("momentum"), ""])
        for ci in range(1, 8):
            c = ws.cell(row=row, column=ci)
            c.fill = PatternFill("solid", fgColor=fill_c)
            c.border = BORDER; c.alignment = _center()
        ws.row_dimensions[row].height = 18

    # Gap row
    ws.append([])

    # Section 2: Stock picks in top sectors
    hdr_row = ws.max_row + 1
    ws.append(["LEADING SECTOR STOCK PICKS"])
    ws.merge_cells(f"A{hdr_row}:H{hdr_row}")
    ws[f"A{hdr_row}"].fill = _hdr_fill("hdr_teal")
    ws[f"A{hdr_row}"].font = _hdr_font(12)
    ws[f"A{hdr_row}"].alignment = _left()
    ws.row_dimensions[hdr_row].height = 24

    hdr2 = ws.max_row + 1
    ws.append(["Symbol","Sector","Setup","Entry","Stop","Target","R:R",
               "Ret30D%","RSI","ADV(Cr)"])
    _style_header_row(ws, hdr2, "hdr_teal")

    for s in sector_picks[:30]:
        ws.append([s.get("symbol"), s.get("sector"), s.get("setup"),
                   s.get("entry"), s.get("stop"), s.get("target"), s.get("rr"),
                   s.get("ret_30d"), s.get("rsi"), s.get("adv_cr")])
        r = ws.max_row
        _style_data_rows(ws, r, r, C["teal_light"])

    _set_col_widths(ws, [14,14,30,9,9,9,8, 9,7,8])


def _sheet_charges(wb, portfolio: list):
    ws = wb.create_sheet("💰 Charge Breakdown")
    if not portfolio:
        ws.append(["No trades"]); return

    cols = ["Symbol","Qty","Investment (₹)","Brokerage (₹)","STT (₹)",
            "Exch (₹)","GST (₹)","Stamp (₹)","SEBI (₹)",
            "Total Charges (₹)","Charges %","Gross R:R","Net R:R"]
    ws.append(cols)
    _style_header_row(ws, 1, "hdr_orange")

    for t in portfolio:
        chg = t.get("charges_breakdown", {})
        ws.append([
            t.get("symbol"), t.get("qty"), t.get("investment"),
            chg.get("brokerage"), chg.get("stt"), chg.get("exch"),
            chg.get("gst"), chg.get("stamp"), chg.get("sebi"),
            t.get("charges_win"), t.get("charges_pct"),
            t.get("gross_rr"), t.get("net_rr"),
        ])
    _style_data_rows(ws, 2, len(portfolio) + 1, C["yellow_light"], C["white"])
    _set_col_widths(ws, [12,7,14,14,12,12,12,12,12,16,10,10,10])


def _sheet_risk_summary(wb, risk: dict, portfolio: list, sector_info: list):
    ws = wb.create_sheet("📋 Risk Summary")

    def section(title):
        r = ws.max_row + 1
        ws.append([title, ""])
        ws.merge_cells(f"A{r}:B{r}")
        ws[f"A{r}"].fill = _hdr_fill("hdr_dark")
        ws[f"A{r}"].font = _hdr_font(11)
        ws[f"A{r}"].alignment = _left()
        ws.row_dimensions[r].height = 24

    def row(label, value, good=None):
        r_idx = ws.max_row + 1
        ws.append([label, value])
        is_good = good is True
        is_bad  = good is False
        fg = C["green_light"] if is_good else (C["red_light"] if is_bad else C["gray_light"])
        for col in ["A", "B"]:
            c = ws[f"{col}{r_idx}"]
            c.fill = PatternFill("solid", fgColor=fg)
            c.font = Font(bold=(col == "A"), size=10)
            c.alignment = _left() if col == "A" else Alignment(horizontal="right", vertical="center")
            c.border = BORDER
        ws.row_dimensions[r_idx].height = 20

    section("CAPITAL ALLOCATION")
    row("Total Capital",              f"₹{CAPITAL:,.0f}")
    row("Deployed",                   f"₹{risk.get('total_invest',0):,.0f}  ({risk.get('deploy_pct',0):.1f}%)")
    row("In Reserve",                 f"₹{risk.get('capital_reserve',0):,.0f}  ({100-risk.get('deploy_pct',0):.1f}%)")
    row("Number of Trades",           risk.get("num_trades", 0))
    ws.append([])

    section("RISK & REWARD")
    row("Total Risk (all stop out)",  f"₹{risk.get('total_risk',0):,.0f}  ({risk.get('total_risk_pct',0):.2f}%)", good=risk.get('total_risk_pct',0) <= 8)
    row("Total Reward (all target)",  f"₹{risk.get('total_reward',0):,.0f}", good=True)
    row("Portfolio Gross R:R",        risk.get("portfolio_rr", 0))
    row("Average Net R:R per trade",  risk.get("avg_net_rr", 0), good=risk.get("avg_net_rr",0) >= 1.5)
    row("Total Charges (win scen.)",  f"₹{risk.get('total_charges',0):,.0f}")
    row("Net Profit (all targets hit)", f"₹{risk.get('net_profit_all',0):,.0f}", good=True)
    row("Net Loss (all stops hit)",   f"₹{risk.get('net_loss_all',0):,.0f}", good=False)
    row("Break-even Win Rate",        f"{risk.get('be_win_rate',50):.1f}%")
    ws.append([])

    section("BREAK-EVEN ANALYSIS")
    net_p = risk.get("net_profit_all", 0)
    net_l = risk.get("net_loss_all", 0)
    for wr in [35, 40, 45, 50, 55, 60]:
        exp = round(wr/100 * net_p - (1-wr/100) * net_l, 0)
        row(f"If Win Rate = {wr}%", f"₹{exp:,.0f}", good=exp > 0)
    ws.append([])

    section("SECTOR EXPOSURE")
    for sec, pct in (risk.get("sector_exposure") or {}).items():
        row(sec, f"{pct}% of capital", good=pct <= 25)
    ws.append([])

    section("WARNINGS & NOTES")
    warnings = risk.get("warnings", [])
    if warnings:
        for w in warnings:
            row("⚠ WARNING", w, good=False)
    else:
        row("✓ All risk checks passed", "Portfolio within guidelines", good=True)

    row("Broker", "Zerodha NRE — ₹200/order (delivery CNC)")
    row("F&O / Intraday", "NOT permitted on NRE account")
    row("Risk per trade", "1% of total capital (fixed)")
    row("Position cap", "15% max / 2% min per trade")
    row("Sector cap", "Max 3 stocks per sector")

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 50


# ── Main save function ────────────────────────────────────────────────────────

def save_report(portfolio: list, sector_info: list, risk_summary: dict,
                all_signals_scored: list, all_raw_signals: dict) -> str:
    """Generate and save the full Excel report. Returns file path."""

    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename  = f"05_institutional_strategies_{timestamp}.xlsx"
    filepath  = os.path.join(RESULTS_DIR, filename)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        # Create a dummy sheet (will be deleted after)
        pd.DataFrame().to_excel(writer, sheet_name="_tmp")
        wb = writer.book

        # ── Build sheets ──────────────────────────────────────────────
        _sheet_portfolio(wb, portfolio)
        _sheet_all_signals(wb, all_signals_scored)

        # Mean Reversion
        mr_cols = [
            ("Symbol",      "symbol",    12),
            ("Setup",       "setup",     40),
            ("Entry",       "entry",      9),
            ("Stop",        "stop",       9),
            ("Target",      "target",     9),
            ("R:R",         "rr",         7),
            ("RSI",         "rsi",        6),
            ("Z-Score",     "zscore",     9),
            ("BB %",        "bb_pct",     8),
            ("Vol Ratio",   "vol_ratio",  9),
            ("ADV (Cr)",    "adv_cr",     9),
            ("RSI Div",     "rsi_div",    9),
            ("Conviction",  "conviction", 10),
        ]
        _sheet_strategy(wb, "📉 Mean Reversion",
                        all_raw_signals.get("mean_reversion", []),
                        mr_cols, "hdr_blue", C["blue_light"])

        # Momentum
        mom_cols = [
            ("Symbol",       "symbol",      12),
            ("Setup",        "setup",       40),
            ("Entry",        "entry",        9),
            ("Stop",         "stop",         9),
            ("Target",       "target",       9),
            ("R:R",          "rr",           7),
            ("ADX",          "adx",          7),
            ("Vol Ratio",    "vol_ratio",    9),
            ("EMA Aligned",  "ema_aligned",  10),
            ("12-1 Mom %",   "ret_12_1",     10),
            ("1M Ret %",     "ret_1m",       9),
            ("52W High %",   "pct_52w_high", 10),
            ("ADV (Cr)",     "adv_cr",        9),
            ("Conviction",   "conviction",   10),
        ]
        _sheet_strategy(wb, "🚀 Momentum",
                        all_raw_signals.get("momentum", []),
                        mom_cols, "hdr_green", C["green_light"])

        # Factor Model
        factor_cols = [
            ("Rank",         "rank",        6),
            ("Symbol",       "symbol",      12),
            ("Name",         "name",        24),
            ("Sector",       "sector",      14),
            ("Score",        "composite",    9),
            ("12-1 Mom %",   "mom_12_1",    10),
            ("ROE %",        "roe",          8),
            ("P/B",          "pb",           7),
            ("P/E",          "pe",           7),
            ("EPS Gr %",     "eps_growth",   9),
            ("Vol 1Y %",     "vol_1y",       9),
            ("Entry",        "entry",        9),
            ("Stop",         "stop",         9),
            ("Target",       "target",       9),
            ("R:R",          "rr",           7),
            ("RSI",          "rsi",          6),
            ("EMA Trend",    "ema_trend",    10),
        ]
        _sheet_strategy(wb, "📐 Factor Model",
                        all_raw_signals.get("factor_model", []),
                        factor_cols, "hdr_purple", C["purple_light"])

        # Pairs Trading
        pairs_cols = [
            ("Long Leg",     "long_leg",    12),
            ("Short Leg",    "short_leg",   12),
            ("Pair",         "pair",        24),
            ("Z-Score",      "z_score",      9),
            ("Correlation",  "correlation",  10),
            ("Hedge Ratio",  "hedge_ratio",  10),
            ("Half-Life(d)", "half_life_days",12),
            ("Entry (Long)", "entry",         9),
            ("Stop",         "stop",          9),
            ("Target",       "target",        9),
            ("R:R",          "rr",            7),
            ("Note",         "note",         36),
            ("Conviction",   "conviction",   10),
        ]
        _sheet_strategy(wb, "↔ Pairs Trading",
                        all_raw_signals.get("pairs_trading", []),
                        pairs_cols, "hdr_teal", C["teal_light"])

        # Sector Rotation
        _sheet_sector_rotation(wb,
                               all_raw_signals.get("sector_rotation_info", []),
                               all_raw_signals.get("sector_rotation_picks", []))

        # PEAD
        pead_cols = [
            ("Symbol",      "symbol",     12),
            ("Setup",       "setup",      36),
            ("Gap %",       "gap_pct",     9),
            ("Vol Surge",   "vol_surge",   9),
            ("Days Ago",    "days_ago",    8),
            ("Entry",       "entry",       9),
            ("Stop",        "stop",        9),
            ("Gap Fill",    "gap_fill",    9),
            ("Target",      "target",      9),
            ("R:R",         "rr",          7),
            ("Hold Days",   "hold_days",   8),
            ("RSI",         "rsi",         6),
            ("ADV (Cr)",    "adv_cr",      9),
            ("Conviction",  "conviction",  10),
        ]
        _sheet_strategy(wb, "📈 PEAD",
                        all_raw_signals.get("pead", []),
                        pead_cols, "hdr_orange", C["yellow_light"])

        # Charges & Risk
        _sheet_charges(wb, portfolio)
        _sheet_risk_summary(wb, risk_summary, portfolio, sector_info)

        # Remove placeholder sheet and reorder
        if "_tmp" in wb.sheetnames:
            del wb["_tmp"]

        # Desired order
        desired_order = [
            "📊 Portfolio Plan", "🔎 All Signals",
            "📉 Mean Reversion", "🚀 Momentum", "📐 Factor Model",
            "↔ Pairs Trading", "🔄 Sector Rotation", "📈 PEAD",
            "💰 Charge Breakdown", "📋 Risk Summary",
        ]
        for i, name in enumerate(desired_order):
            if name in wb.sheetnames:
                wb.move_sheet(name, offset=i - wb.index(wb[name]))

    print(f"\n  [OK] Report saved --> {filepath}")
    return filepath
