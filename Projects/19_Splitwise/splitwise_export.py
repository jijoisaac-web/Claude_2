"""
Splitwise Excel Exporter
========================
Exports to Excel:
  - Summary   : totals for unsettled groups
  - Groups    : only groups with an outstanding balance
  - Expenses  : last 6 months, only from unsettled groups, with Payment/Expense type

Usage:
  pip install requests openpyxl
  python splitwise_export.py
"""

import webbrowser
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import urlparse, parse_qs
from datetime import datetime, timedelta

# ── Credentials ───────────────────────────────────────────────────────────────
CONSUMER_KEY    = "2N61qBp6YFaKUIO1ksnIQkNhuvqjCOzYykwReofR"
CONSUMER_SECRET = "IUurkcgCRyVroPdbptAJvAF5xsXDkT9Hssj6eHch"
REDIRECT_URI    = "http://localhost:8080/callback"
OUTPUT_FILE     = r"C:\Users\Ansa\Claude\Projects\19_Splitwise\Splitwise_Unsettled_6Months_{}.xlsx".format(datetime.now().strftime("%Y%m%d_%H%M%S"))
# ─────────────────────────────────────────────────────────────────────────────

# ── Colours ───────────────────────────────────────────────────────────────────
GREEN_DARK   = "1B6C3E"
GREEN_LIGHT  = "E8F5E9"
RED_LIGHT    = "FFEBEE"
BLUE_LIGHT   = "E3F2FD"
GREY_HEADER  = "37474F"
WHITE        = "FFFFFF"
# ─────────────────────────────────────────────────────────────────────────────

auth_code = None

class CallbackHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        global auth_code
        params = parse_qs(urlparse(self.path).query)
        if "code" in params:
            auth_code = params["code"][0]
            self.send_response(200)
            self.end_headers()
            self.wfile.write(b"<h2 style='font-family:sans-serif;color:green'>Authorization successful! You can close this tab.</h2>")
        else:
            self.send_response(400)
            self.end_headers()
            self.wfile.write(b"<h2>No code received.</h2>")
    def log_message(self, *args):
        pass


def get_access_token():
    auth_url = (
        f"https://secure.splitwise.com/oauth/authorize"
        f"?client_id={CONSUMER_KEY}&response_type=code&redirect_uri={REDIRECT_URI}"
    )
    print("\n🔐 Opening Splitwise in your browser for authorization...")
    webbrowser.open(auth_url)
    server = HTTPServer(("localhost", 8080), CallbackHandler)
    print("⏳ Waiting for you to approve in the browser...")
    server.handle_request()
    if not auth_code:
        print("❌ Authorization failed.")
        return None
    resp = requests.post(
        "https://secure.splitwise.com/oauth/token",
        data={
            "grant_type":    "authorization_code",
            "client_id":     CONSUMER_KEY,
            "client_secret": CONSUMER_SECRET,
            "code":          auth_code,
            "redirect_uri":  REDIRECT_URI,
        }
    )
    token_data = resp.json()
    if "access_token" not in token_data:
        print(f"❌ Token exchange failed: {token_data}")
        return None
    print("✅ Authorized!\n")
    return token_data["access_token"]


def sw_get(endpoint, token):
    resp = requests.get(
        f"https://secure.splitwise.com/api/v3.0/{endpoint}",
        headers={"Authorization": f"Bearer {token}"}
    )
    return resp.json()


def clean_name(user):
    if isinstance(user, dict):
        first = user.get("first_name") or ""
        last  = user.get("last_name") or ""
        name  = f"{first} {last}".strip()
        return name if name else user.get("email", "Unknown")
    return str(user)


def get_unsettled_groups(groups, my_id):
    """Return groups where the user has a non-zero balance."""
    unsettled = []
    for g in groups:
        for m in g.get("members", []):
            if str(m.get("id")) == str(my_id):
                for b in m.get("balance", []):
                    if float(b.get("amount", 0)) != 0:
                        unsettled.append(g)
                        break
    return unsettled


def fetch_expenses_last_6_months(token, group_ids):
    """Fetch last 6 months of expenses only for given group IDs."""
    dated_after = (datetime.now() - timedelta(days=1095)).strftime("%Y-%m-%dT00:00:00Z")
    all_expenses = []
    for gid in group_ids:
        offset = 0
        limit  = 100
        while True:
            data  = sw_get(
                f"get_expenses?group_id={gid}&limit={limit}&offset={offset}&dated_after={dated_after}",
                token
            )
            batch = data.get("expenses", [])
            if not batch:
                break
            all_expenses.extend(batch)
            offset += limit
            if len(batch) < limit:
                break
    # Also fetch non-group expenses (group_id=0 means no group)
    offset = 0
    while True:
        data  = sw_get(
            f"get_expenses?limit=100&offset={offset}&dated_after={dated_after}",
            token
        )
        batch = [e for e in data.get("expenses", []) if not e.get("group_id")]
        if not batch:
            break
        all_expenses.extend(batch)
        offset += 100
        if len(data.get("expenses", [])) < 100:
            break

    # Deduplicate by expense id
    seen = set()
    unique = []
    for e in all_expenses:
        if e["id"] not in seen:
            seen.add(e["id"])
            unique.append(e)
    return unique


# ── Excel helpers ─────────────────────────────────────────────────────────────

def header_style(cell, bg=GREY_HEADER, fg=WHITE):
    cell.font      = Font(bold=True, color=fg, size=11)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = Border(
        bottom=Side(style="medium", color="AAAAAA")
    )

def money_color(cell, amount):
    if amount > 0:
        cell.fill = PatternFill("solid", fgColor=GREEN_LIGHT)
        cell.font = Font(color="1B6C3E", bold=True)
    elif amount < 0:
        cell.fill = PatternFill("solid", fgColor=RED_LIGHT)
        cell.font = Font(color="C62828", bold=True)

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def alt_row(ws, ri, num_cols):
    if ri % 2 == 0:
        for ci in range(1, num_cols + 1):
            cell = ws.cell(row=ri, column=ci)
            if not cell.fill or cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                cell.fill = PatternFill("solid", fgColor="F9F9F9")


# ── Sheet builders ─────────────────────────────────────────────────────────────

def build_summary(wb, me, unsettled_groups, my_id, expenses):
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    title = ws["A1"]
    title.value     = "Splitwise — Unsettled Groups (Last 6 Months)"
    title.font      = Font(bold=True, size=15, color=GREEN_DARK)
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 34

    ws["A2"] = f"Account: {clean_name(me)}  |  {me.get('email', '')}"
    ws["A2"].font = Font(italic=True, color="555555", size=10)
    ws["A3"] = f"Exported: {datetime.now().strftime('%d %B %Y, %I:%M %p')}  |  Period: {(datetime.now()-timedelta(days=1095)).strftime('%d %b %Y')} – {datetime.now().strftime('%d %b %Y')}"
    ws["A3"].font = Font(italic=True, color="555555", size=10)
    ws.row_dimensions[3].height = 16

    # Per-group balance summary
    ws.row_dimensions[5].height = 22
    headers = ["Group Name", "Currency", "Your Balance", "Status", "Transactions (6mo)"]
    for ci, h in enumerate(headers, 1):
        header_style(ws.cell(row=5, column=ci, value=h))

    # Count expenses per group
    exp_count = {}
    for e in expenses:
        if not e.get("deleted_at"):
            gid = str(e.get("group_id") or "non-group")
            exp_count[gid] = exp_count.get(gid, 0) + 1

    ri = 6
    total_owed_to_me = 0.0
    total_i_owe      = 0.0
    for g in unsettled_groups:
        gid  = str(g.get("id"))
        name = g.get("name", "")
        for m in g.get("members", []):
            if str(m.get("id")) == str(my_id):
                for b in m.get("balance", []):
                    amt = float(b.get("amount", 0))
                    cur = b.get("currency_code", "")
                    if amt == 0:
                        continue
                    ws.cell(row=ri, column=1, value=name).font = Font(bold=True)
                    ws.cell(row=ri, column=2, value=cur)
                    c = ws.cell(row=ri, column=3, value=round(amt, 2))
                    c.number_format = '#,##0.00'
                    money_color(c, amt)
                    status = "Owed to you" if amt > 0 else "You owe"
                    ws.cell(row=ri, column=4, value=status)
                    ws.cell(row=ri, column=5, value=exp_count.get(gid, 0))
                    alt_row(ws, ri, 5)
                    if amt > 0:
                        total_owed_to_me += amt
                    else:
                        total_i_owe += abs(amt)
                    ri += 1

    # Totals row
    ri += 1
    ws.cell(row=ri, column=1, value="TOTAL").font = Font(bold=True, size=11)
    c_owed = ws.cell(row=ri, column=3, value=round(total_owed_to_me - total_i_owe, 2))
    c_owed.number_format = '#,##0.00'
    c_owed.font = Font(bold=True, size=11)
    money_color(c_owed, total_owed_to_me - total_i_owe)
    ws.cell(row=ri, column=4, value=f"Owed to you: {round(total_owed_to_me,2)}  |  You owe: {round(total_i_owe,2)}")

    for col, width in zip([1,2,3,4,5], [30, 12, 16, 16, 22]):
        set_col_width(ws, col, width)


def build_expenses(wb, me, unsettled_groups, expenses):
    ws = wb.create_sheet("Expenses")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    headers = ["Date", "Type", "Group", "Description", "Currency",
               "Total Cost", "Paid By", "Your Share", "Your Net"]
    for ci, h in enumerate(headers, 1):
        header_style(ws.cell(row=1, column=ci, value=h))
    ws.row_dimensions[1].height = 22

    group_map = {str(g.get("id")): g.get("name", "Unknown") for g in unsettled_groups}
    my_id = str(me.get("id"))

    ri = 2
    active = sorted(
        [e for e in expenses if not e.get("deleted_at")],
        key=lambda x: x.get("date", ""),
        reverse=True
    )

    for e in active:
        date       = e.get("date", "")[:10]
        desc       = e.get("description", "")
        group_id   = str(e.get("group_id") or "")
        group_name = group_map.get(group_id, "Non-group")
        cur        = e.get("currency_code", "")
        cost       = float(e.get("cost", 0))
        is_payment = e.get("payment", False)
        txn_type   = "Payment" if is_payment else "Expense"

        users      = e.get("users", [])
        payer_name = ""
        my_paid    = 0.0
        my_owed    = 0.0
        for u in users:
            u_info = u.get("user") or {}
            paid_share = float(u.get("paid_share", 0))
            owed_share = float(u.get("owed_share", 0))
            if paid_share > 0 and not payer_name:
                payer_name = clean_name(u_info)
            if str(u_info.get("id")) == my_id:
                my_paid = paid_share
                my_owed = owed_share

        # Skip transactions where I have no involvement
        if my_paid == 0.0 and my_owed == 0.0:
            continue

        my_net = my_paid - my_owed

        ws.cell(row=ri, column=1, value=date)

        # Type cell — colour-coded
        c_type = ws.cell(row=ri, column=2, value=txn_type)
        if is_payment:
            c_type.fill = PatternFill("solid", fgColor=BLUE_LIGHT)
            c_type.font = Font(color="1565C0", bold=True)
        else:
            c_type.font = Font(color="555555")

        ws.cell(row=ri, column=3, value=group_name)
        ws.cell(row=ri, column=4, value=desc)
        ws.cell(row=ri, column=5, value=cur)

        c_cost = ws.cell(row=ri, column=6, value=round(cost, 2))
        c_cost.number_format = '#,##0.00'

        ws.cell(row=ri, column=7, value=payer_name)

        c_share = ws.cell(row=ri, column=8, value=round(my_owed, 2))
        c_share.number_format = '#,##0.00'

        c_net = ws.cell(row=ri, column=9, value=round(my_net, 2))
        c_net.number_format = '#,##0.00'
        money_color(c_net, my_net)

        alt_row(ws, ri, 9)
        ri += 1

    for col, width in zip(range(1, 10), [13, 11, 22, 35, 10, 13, 22, 13, 13]):
        set_col_width(ws, col, width)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    token = get_access_token()
    if not token:
        return

    print("📡 Fetching your data from Splitwise...")
    me     = sw_get("get_current_user", token).get("user", {})
    groups = sw_get("get_groups", token).get("groups", [])
    my_id  = str(me.get("id"))

    # Filter to unsettled groups only
    unsettled = get_unsettled_groups(groups, my_id)
    unsettled_ids = [str(g.get("id")) for g in unsettled]
    print(f"   Found {len(unsettled)} unsettled groups out of {len(groups)} total")

    # Fetch last 6 months of expenses for those groups
    print("📥 Fetching last 6 months of expenses", end="", flush=True)
    expenses = fetch_expenses_last_6_months(token, unsettled_ids)
    active   = [e for e in expenses if not e.get("deleted_at")]
    print(f" {len(active)} transactions found\n")

    print("📊 Building Excel file...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_summary(wb, me, unsettled, my_id, active)
    build_expenses(wb, me, unsettled, active)

    wb.save(OUTPUT_FILE)
    print(f"\n✅ Done! Saved: {OUTPUT_FILE}")
    print(f"   Sheets: Summary | Expenses")
    print(f"   {len(unsettled)} unsettled groups  |  {len(active)} transactions")


if __name__ == "__main__":
    main()
